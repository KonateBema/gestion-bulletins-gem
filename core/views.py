# core/views.py

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, FileResponse
from django.views.decorators.csrf import csrf_protect
from .forms import EtudiantForm,ClasseForm,MatiereForm,AffectationForm,NoteForm
from django.shortcuts import get_object_or_404
from django.core.paginator import Paginator
from django.db.models import Q
from .services import calcul_moyenne_etudiant
from openpyxl import load_workbook
from .models import Filierebts
from django.contrib import messages
from .models import Salle
from .models import SaisieNotesBTS
from .models import ( Classe,Niveau)
from lmd.models import EtudiantLMD,FiliereLMD
# from .models import NoteBTS
from .models import (Etudiant, Professeur, Matiere, Note,AffectationMatiere, Inscription, Profile)
from .forms import UserRegisterForm
from .utils import generate_matricule
from .services import (mention,)
from .pdf_service import generate_bulletin_pdf
from datetime import datetime
from core.decorators import role_required
from django.http import JsonResponse
from .models import GrandeUnite
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER,TA_LEFT
from reportlab.lib.units import cm

# =========================
# 🔐 LOGIN
# =========================
@csrf_protect
def login_viewAAAA(request):

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            profile = Profile.objects.filter(user=user).first()

            if not profile:
                return render(request, "login.html", {
                    "error": "Profil utilisateur introuvable"
                })

            if profile.role == "ADMIN":
                return redirect("dashboard_admin")

            elif profile.role == "PROF":
                return redirect("dashboard_prof")

            else:
                return redirect("dashboard_etudiant")

        return render(request, "login.html", {
            "error": "Identifiants incorrects"
        })

    return render(request, "login.html")

from django.views.decorators.csrf import csrf_protect


# ==========================================================
# 🔐 LOGIN
# ==========================================================

@csrf_protect
def login_view(request):

    # ==========================================================
    # UTILISATEUR DÉJÀ CONNECTÉ
    # ==========================================================

    if request.user.is_authenticated:

        profile = Profile.objects.filter(
            user=request.user
        ).first()

        if not profile:

            logout(request)

            return render(
                request,
                "login.html",
                {
                    "error": "Profil utilisateur introuvable."
                }
            )

        # Redirection selon le rôle
        if profile.role == "ADMIN":

            return redirect("dashboard_admin")

        elif profile.role == "GESTIONNAIRE":

            return redirect("dashboard_gestionnaire")

        elif profile.role == "PROF":

            return redirect("dashboard_prof")

        elif profile.role == "ETUDIANT":

            return redirect("dashboard_etudiant")

        else:

            logout(request)

            return render(
                request,
                "login.html",
                {
                    "error": "Rôle utilisateur non reconnu."
                }
            )

    # ==========================================================
    # FORMULAIRE DE CONNEXION
    # ==========================================================

    if request.method == "POST":

        username = request.POST.get(
            "username",
            ""
        ).strip()

        password = request.POST.get(
            "password",
            ""
        )

        # Vérification des champs
        if not username or not password:

            return render(
                request,
                "login.html",
                {
                    "error":
                    "Veuillez renseigner votre identifiant "
                    "et votre mot de passe."
                }
            )

        # ======================================================
        # AUTHENTIFICATION DJANGO
        # ======================================================

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is None:

            return render(
                request,
                "login.html",
                {
                    "error":
                    "Identifiant ou mot de passe incorrect."
                }
            )

        # ======================================================
        # CONNEXION
        # ======================================================

        login(
            request,
            user
        )

        # ======================================================
        # RÉCUPÉRATION DU PROFIL
        # ======================================================

        profile = Profile.objects.filter(
            user=user
        ).first()

        # Aucun profil
        if not profile:

            logout(request)

            return render(
                request,
                "login.html",
                {
                    "error":
                    "Profil utilisateur introuvable."
                }
            )

        # ======================================================
        # REDIRECTION SELON LE RÔLE
        # ======================================================

        if profile.role == "ADMIN":

            return redirect(
                "dashboard_admin"
            )

        elif profile.role == "GESTIONNAIRE":

            return redirect(
                "dashboard_gestionnaire"
            )

        elif profile.role == "PROF":

            return redirect(
                "dashboard_prof"
            )

        elif profile.role == "ETUDIANT":

            return redirect(
                "dashboard_etudiant"
            )

        # ======================================================
        # RÔLE INCONNU
        # ======================================================

        logout(request)

        return render(
            request,
            "login.html",
            {
                "error":
                "Rôle utilisateur non reconnu."
            }
        )

    # ==========================================================
    # AFFICHAGE PAGE LOGIN
    # ==========================================================

    return render(
        request,
        "login.html"
    )


# =========================
# 🚪 LOGOUT
# =========================
def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def dashboard(request):

    filieres_l3 = FiliereLMD.objects.filter(
        niveau_formation="L3"
    )

    filieres_master = FiliereLMD.objects.filter(
        niveau_formation="M1-M2"
    )

    context = {

        # =====================
        # BTS
        # =====================
        "etudiants_count": Etudiant.objects.count(),
        "professeurs_count": Professeur.objects.count(),
        "classes_count": Classe.objects.count(),
        "matieres_count": Matiere.objects.count(),
        "notes_count": Note.objects.count(),


        # =====================
        # LMD
        # =====================
        "l1_count": EtudiantLMD.objects.filter(
            niveau="L1"
        ).count(),

        "l2_count": EtudiantLMD.objects.filter(
            niveau="L2"
        ).count(),

        "l3_count": EtudiantLMD.objects.filter(
            niveau="L3"
        ).count(),

        "master_count": EtudiantLMD.objects.filter(
            niveau__in=["M1", "M2"]
        ).count(),


        # =====================
        # MENU DYNAMIQUE
        # =====================
        "filieres_l3": filieres_l3,

        "filieres_master": filieres_master,

    }


    return render(
        request,
        "dashboard.html",
        context
    )

# =========================
# 🧑‍💼 ADMIN
# =========================
@login_required
def dashboard_admin(request):

    return render(request, "admin_dashboard.html", {
        "etudiants": Etudiant.objects.count(),
        "professeurs": Professeur.objects.count(),
        "classes": Classe.objects.count(),
    })


# =========================
# 👨‍🏫 PROF
# =========================
@login_required
def dashboard_prof(request):

    prof = Professeur.objects.filter(user=request.user).first()

    if not prof:
        return HttpResponse("❌ Profil professeur introuvable")

    matieres = AffectationMatiere.objects.filter(professeur=prof)

    return render(request, "prof_dashboard.html", {
        "matieres": matieres,
    })


# =========================
# 🎓 ETUDIANT
# =========================
@login_required
def dashboard_etudiant(request):

    etudiant = Etudiant.objects.filter(user=request.user).first()

    if not etudiant:
        return HttpResponse("❌ Aucun profil étudiant trouvé")

    notes = Note.objects.filter(etudiant=etudiant)

    return render(request, "etudiant_dashboard.html", {
        "etudiant": etudiant,
        "notes": notes,
    })


# ==========================================================
# 👨‍💼 DASHBOARD GESTIONNAIRE
# ==========================================================

@login_required(login_url="login")
@role_required("GESTIONNAIRE")
def dashboard_gestionnaire(request):

    return render(
        request,
        "gestionnaire_dashboard.html",
        {
            "etudiants": Etudiant.objects.count(),
            "classes": Classe.objects.count(),
            "filieres": Filierebts.objects.count(),
            "matieres": Matiere.objects.count(),
        }
    )



# =========================
# 📝 INSCRIPTION UTILISATEUR
# =========================
def register_user(request):

    if request.method == "POST":

        form = UserRegisterForm(request.POST)

        if form.is_valid():

            user = form.save()

            # 🎓 ETUDIANT
            if user.role == "ETUD":

                etudiant = Etudiant.objects.create(
                    user=user,
                    matricule=generate_matricule("ETU"),
                    date_naissance="2000-01-01",
                    sexe="M",
                    telephone="00000000",
                    classe=Classe.objects.first()
                )

                Inscription.objects.create(
                    etudiant=etudiant,
                    classe=etudiant.classe,
                    annee="2025-2026"
                )

            # 👨‍🏫 PROF
            elif user.role == "PROF":

                Professeur.objects.create(
                    user=user,
                    matricule=generate_matricule("PROF"),
                    specialite="Non définie",
                    telephone="00000000"
                )

            return redirect('login')

    else:
        form = UserRegisterForm()

    return render(request, 'register.html', {'form': form})

# =========================
# 📊 BULLETIN ETUDIANT
# =========================
@role_required("ADMIN")
@login_required
def bulletin_etudiant(request):

    etudiant = Etudiant.objects.first()  # ou filtre propre

    if not etudiant:
        return HttpResponse("Aucun étudiant trouvé")

    moyenne = calcul_moyenne_etudiant(etudiant)

    return render(request, "bulletin.html", {
        "etudiant": etudiant,
        "moyenne": moyenne,
        "mention": mention(moyenne),
    })


# =========================
# 📄 PDF BULLETIN
# =========================
@role_required("ADMIN", "GESTIONNAIRE")
@login_required(login_url="login")
def etudiant_listPRO(request):

    query = request.GET.get("q", "")
    classe_id = request.GET.get("classe", "")
    filiere_bts_id = request.GET.get("filiere_bts", "")
    niveau = request.GET.get("niveau", "")


    # =========================
    # LISTE ETUDIANTS
    # =========================

    etudiants = Etudiant.objects.select_related(
        "classe",
        "filiere_bts"
    ).order_by("nom", "prenoms")



    # =========================
    # RECHERCHE
    # =========================

    if query:

        etudiants = etudiants.filter(

            Q(matricule__icontains=query) |

            Q(nom__icontains=query) |

            Q(prenoms__icontains=query)

        )



    # =========================
    # FILTRE CLASSE
    # =========================

    if classe_id:

        etudiants = etudiants.filter(
            classe_id=classe_id
        )



    # =========================
    # FILTRE NIVEAU BTS
    # =========================

    if niveau:

        etudiants = etudiants.filter(

            filiere_bts__niveaux__nom=niveau

        ).distinct()



    # =========================
    # FILTRE FILIERE BTS
    # =========================

    if filiere_bts_id:

        etudiants = etudiants.filter(

            filiere_bts_id=filiere_bts_id

        )



    # =========================
    # PAGINATION
    # =========================

    paginator = Paginator(
        etudiants,
        10
    )


    page_number = request.GET.get("page")


    page_obj = paginator.get_page(
        page_number
    )



    return render(
        request,
        "etudiants/list.html",
        {

            "page_obj": page_obj,


            # données filtres

            "classes": Classe.objects.all().order_by("nom"),


            "filieres_bts": Filierebts.objects.all().order_by("nom"),


            "niveau": niveau,


            "classe_selected": classe_id,


            "filiere_selected": filiere_bts_id,


            "query": query,

        }
    )

def etudiant_list(request):

    etudiants = Etudiant.objects.select_related(
        "filiere_bts",
        "classe"
    ).all()

    q = request.GET.get("q", "").strip()
    niveau = request.GET.get("niveau", "").strip()
    filiere_bts = request.GET.get("filiere_bts", "").strip()

    # Recherche
    if q:
        etudiants = etudiants.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(matricule__icontains=q)
        )

    # Filtre BTS 1 / BTS 2
    if niveau:
        etudiants = etudiants.filter(
            classe__niveau__nom=niveau
        )

    # Filtre filière
    if filiere_bts:
        etudiants = etudiants.filter(
            filiere_bts_id=filiere_bts
        )

    filieres_bts = Filierebts.objects.all().order_by("nom")

    paginator = Paginator(etudiants, 20)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "filieres_bts": filieres_bts,
    }

    return render(
        request,
        "etudiants/list.html",
        context
    )


def etudiant_create(request):

    if request.method == "POST":

        form = EtudiantForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect('etudiant_list')

    else:
        form = EtudiantForm()

    filieres_bts = Filierebts.objects.all()

    return render(request, 'etudiants/form.html', {
        'form': form,
        'filieres_bts': filieres_bts
    })


def etudiants_par_salle(request):

    salles = Salle.objects.prefetch_related(
        'classe_set__etudiants'
    )

    return render(
        request,
        'etudiants/par_salle.html',
        {
            'salles': salles
        }
    )

def etudiant_update(request, id):

    etudiant = get_object_or_404(Etudiant, id=id)

    if request.method == "POST":

        form = EtudiantForm(request.POST, instance=etudiant)

        if form.is_valid():
            form.save()
            return redirect('etudiant_list')

    else:

        form = EtudiantForm(instance=etudiant)

    return render(request, 'etudiants/form.html', {
        'form': form
    })


def etudiant_delete(request, id):
    Etudiant.objects.get(id=id).delete()
    return redirect("etudiant_list")

def classe_list(request):

    # ==========================
    # CREATION CLASSE
    # ==========================
    if request.method == "POST":

        form = ClasseForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect("classe_list")

        else:
            print(form.errors)

    else:
        form = ClasseForm()



    # ==========================
    # FILTRES
    # ==========================
    query = request.GET.get("q", "")
    filiere_bts = request.GET.get("filiere_bts", "")
    niveau = request.GET.get("niveau", "")



    # ==========================
    # LISTE DES CLASSES
    # ==========================
    classes = Classe.objects.select_related(
        "filiere_bts",
        "niveau",
        "salle"
    ).order_by("-id")



    # Recherche par nom
    if query:
        classes = classes.filter(
            nom__icontains=query
        )


    # Filtre filière BTS
    if filiere_bts:
        classes = classes.filter(
            filiere_bts_id=filiere_bts
        )


    # Filtre niveau
    if niveau:
        classes = classes.filter(
            niveau_id=niveau
        )



    # ==========================
    # PAGINATION
    # ==========================
    paginator = Paginator(
        classes,
        10
    )

    page_number = request.GET.get(
        "page"
    )

    page_obj = paginator.get_page(
        page_number
    )



    # ==========================
    # DONNEES POUR SELECTS
    # ==========================
    filieres = Filierebts.objects.all()

    niveaux = Niveau.objects.all()  # noqa: F821



    return render(
        request,
        "classes/list.html",
        {
            "page_obj": page_obj,
            "form": form,
            "filieres": filieres,
            "niveaux": niveaux,
            "query": query,
            "filiere_selected": filiere_bts,
            "niveau_selected": niveau,
        }
    )


def classe_create(request):
    if request.method == "POST":
        form = ClasseForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect('classe_list')
    else:
        form = ClasseForm()

    return render(request, 'classes/form.html', {
        'form': form
    })


def matiere_listAAA(request):

    filieres = Filierebts.objects.prefetch_related(
        "matiere_set"
    ).order_by("nom")


    q = request.GET.get("q")


    if q:

        filieres = filieres.filter(
            matiere__libelle__icontains=q
        ).distinct()



    context = {

        "filieres": filieres,

        "total_matieres": Matiere.objects.count(),

    }


    return render(
        request,
        "matieres/list.html",
        context
    )

from django.core.paginator import Paginator
from django.db.models import Q

def matiere_list(request):

    q = request.GET.get("q", "")
    filiere_id = request.GET.get("filiere_bts", "")

    matieres = Matiere.objects.select_related(
        "filiere_bts"
    ).order_by(
        "filiere_bts__nom",
        "libelle"
    )

    if q:
        matieres = matieres.filter(
            Q(libelle__icontains=q) |
            Q(code__icontains=q)
        )

    if filiere_id:
        matieres = matieres.filter(
            filiere_bts_id=filiere_id
        )

    paginator = Paginator(matieres, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "total_matieres": matieres.count(),
        "filiere_list": Filierebts.objects.order_by("nom"),
        "filiere_selectionnee": filiere_id,
    }

    return render(
        request,
        "matieres/list.html",
        context, 
    )


def matiere_create(request):
    form = MatiereForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect("matiere_list")
    return render(request, "matieres/form.html", {"form": form})


def affectation_list(request):
    return render(request, "affectations/list.html", {
        "affectations": Affectation.objects.select_related(  # noqa: F821
            "professeur", "matiere", "classe"
        )
    })


def affectation_create(request):
    form = AffectationForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect("affectation_list")

    return render(request, "affectations/form.html", {
        "form": form,
        "title": "Affecter professeur"
    })


def affectation_delete(request, id):
    Affectation.objects.get(id=id).delete()  # noqa: F821
    return redirect("affectation_list")



def note_list(request):

    notes = Note.objects.select_related(
        "etudiant",
        "matiere"
    ).all().order_by("-id")

    etudiant_id = request.GET.get("etudiant")
    matiere_id = request.GET.get("matiere")
    semestre = request.GET.get("semestre")

    if etudiant_id:
        notes = notes.filter(etudiant_id=etudiant_id)

    if matiere_id:
        notes = notes.filter(matiere_id=matiere_id)

    if semestre:
        notes = notes.filter(semestre=semestre)

    paginator = Paginator(notes, 10)  # 10 lignes par page

    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "notes/list.html", {
        "notes": page_obj,
        "page_obj": page_obj,
        "etudiants": Etudiant.objects.all(),
        "matieres": Matiere.objects.all(),
    })

def note_create(request):

    form = NoteForm(request.POST or None)

    if request.method == "POST":

        if form.is_valid():
            note = form.save(commit=False)

            # DEBUG OPTIONNEL
            print("✔ Note enregistrée")

            note.save()
            return redirect("note_list")

        else:
            print(form.errors)

    return render(request, "notes/form.html", {
        "form": form,
        "title": "Ajouter note"
    })
    
def note_update(request, id):
    note = Note.objects.get(id=id)
    form = NoteForm(request.POST or None, instance=note)

    if form.is_valid():
        form.save()
        return redirect("note_list")

    return render(request, "notes/form.html", {
        "form": form,
        "title": "Modifier note"
    })


def note_delete(request, id):
    Note.objects.get(id=id).delete()
    return redirect("note_list")

def moyenne_etudiant(etudiant):
    notes = Note.objects.filter(etudiant=etudiant)

    if not notes:
        return 0

    total = sum(n.moyenne for n in notes)
    return total / notes.count()


def classe_edit(request, pk):
    classe = Classe.objects.get(pk=pk)
    form = ClasseForm(request.POST or None, instance=classe)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            return redirect('classe_list')

    return render(request, 'classes/form.html', {
        'form': form
    })

def classe_delete(request, pk):
    classe = get_object_or_404(Classe, pk=pk)
    classe.delete()
    return redirect('classe_list')

def matiere_update(request, id):
    matiere = get_object_or_404(Matiere, id=id)
    form = MatiereForm(request.POST or None, instance=matiere)

    if form.is_valid():
        form.save()
        return redirect('matiere_list')

    return render(request, 'matieres/form.html', {'form': form})

def matiere_delete(request, id):
    matiere = get_object_or_404(Matiere, id=id)
    matiere.delete()
    return redirect('matiere_list')


def download_bulletin_pdf(request, etudiant_id, classe_id, semestre):

    etudiant = get_object_or_404(
        Etudiant,
        id=etudiant_id
    )

    classe = get_object_or_404(
        Classe,
        id=classe_id
    )

    file_path = generate_bulletin_pdf(
        etudiant=etudiant,
        classe=classe,
        semestre=semestre
    )

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename=f"bulletin_S{semestre}_{etudiant.matricule}.pdf"
    )

def bulletin_classe(request, classe_id):

    classe = Classe.objects.get(id=classe_id)

    data = classement(classe)  # noqa: F821

    return render(request, "bulletin_classe.html", {
        "classe": classe,
        "data": data
    })

@login_required(login_url="login")
def bulletin_list(request):

    etudiants = Etudiant.objects.select_related("classe").all()

    # 🔎 Filtres GET
    matricule = request.GET.get("matricule")
    telephone = request.GET.get("telephone")
    filiere = request.GET.get("filiere")
    classe = request.GET.get("classe")

    # 🔽 Filtrage dynamique
    if matricule:
        etudiants = etudiants.filter(matricule__icontains=matricule)

    if telephone:
        etudiants = etudiants.filter(telephone__icontains=telephone)

    if filiere:
        etudiants = etudiants.filter(filiere__icontains=filiere)

    if classe:
        etudiants = etudiants.filter(classe_id=classe)

    # 📄 PAGINATION
    paginator = Paginator(etudiants, 10)  # 10 étudiants par page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "bulletins/list.html", {
        "etudiants": page_obj,
        "page_obj": page_obj,
        "classes": Classe.objects.all(),
    })
 


def liste_filieres_bts(request):
    filieres = Filierebts.objects.all().order_by('nom')

    return render(request, 'bts/liste_filieres_bts.html', {
        'filieres': filieres
    })

def ajouter_filiere_btsGGG(request):
    if request.method == "POST":
        nom = request.POST.get("nom")

        Filierebts.objects.create(
            nom=nom
        )

        messages.success(request, "Filière BTS ajoutée avec succès.")
        return redirect('liste_filieres_bts')

    return render(request, 'bts/ajouter_filiere_bts.html')


def ajouter_filiere_bts(request):

    if request.method == "POST":

        nom = request.POST.get("nom")
        niveaux_ids = request.POST.getlist("niveaux")

        # Création de la filière
        filiere = Filierebts.objects.create(
            nom=nom
        )

        # Association des niveaux sélectionnés
        filiere.niveaux.set(niveaux_ids)

        messages.success(
            request,
            "Filière BTS ajoutée avec succès."
        )

        return redirect("liste_filieres_bts")

    niveaux = Niveau.objects.all()

    return render(
        request,
        "bts/ajouter_filiere_bts.html",
        {
            "niveaux": niveaux
        }
    )

def modifier_filiere_bts(request, pk):

    filiere = get_object_or_404(Filierebts, pk=pk)
    niveau = filiere.niveaux.first()

    if request.method == "POST":

        filiere.nom = request.POST.get("nom")
        filiere.save()

        nom_niveau = request.POST.get("niveau")

        niveau, _ = Niveau.objects.get_or_create(
            nom=nom_niveau
        )

        filiere.niveaux.set([niveau])

        messages.success(request, "Filière modifiée avec succès.")
        return redirect("liste_filieres_bts")

    return render(
        request,
        "bts/modifier_filiere_bts.html",
        {
            "filiere": filiere,
            "niveau": niveau,
        }
    )


def supprimer_filiere_bts(request, pk):
    filiere = get_object_or_404(Filierebts, pk=pk)

    filiere.delete()

    messages.success(request, "Filière BTS supprimée.")
    return redirect('liste_filieres_bts')

def salle_list(request):
    salles = Salle.objects.all()
    return render(request, 'salles/salle_list.html', {
        'salles': salles
    })

def salle_create(request):
    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")
        capacite = request.POST.get("capacite")

        Salle.objects.create(
            code=code,
            nom=nom,
            capacite=capacite
        )

        messages.success(request, "Salle ajoutée avec succès")
        return redirect('salle_list')

    return render(request, 'salles/salle_form.html')

def salle_edit(request, pk):
    salle = get_object_or_404(Salle, pk=pk)

    if request.method == "POST":
        salle.code = request.POST.get("code")
        salle.nom = request.POST.get("nom")
        salle.capacite = request.POST.get("capacite")
        salle.save()

        messages.success(request, "Salle modifiée avec succès")
        return redirect('salle_list')

    return render(request, 'salles/salle_form.html', {
        'salle': salle
    })

def salle_delete(request, pk):
    salle = get_object_or_404(Salle, pk=pk)
    salle.delete()

    messages.success(request, "Salle supprimée")
    return redirect('salle_list')


def saisie_note_groupee(request):

    classes = Classe.objects.select_related(
        "filiere_bts",
        "niveau",
        "salle"
    )

    matieres = Matiere.objects.all()

    etudiants = []
    notes_existantes = {}

    classe_id = request.GET.get("classe")
    matiere_id = request.GET.get("matiere")
    semestre = request.GET.get("semestre")


    # ==========================
    # ENREGISTREMENT DES NOTES
    # ==========================
    if request.method == "POST":

        classe_id = request.POST.get("classe")
        matiere_id = request.POST.get("matiere")
        semestre = request.POST.get("semestre")


        classe = Classe.objects.get(id=classe_id)

        etudiants = Etudiant.objects.filter(
            classe=classe
        )


        # créer ou récupérer une saisie
        saisie, created = SaisieNotesBTS.objects.get_or_create(
            classe=classe,
            matiere_id=matiere_id,
            semestre=semestre
        )


        for etudiant in etudiants:

            cc = float(
                request.POST.get(f"cc_{etudiant.id}") or 0
            )
            
            devoir = float(
               request.POST.get(f"devoir_{etudiant.id}") or 0
            )

            examen = float(
                request.POST.get(f"examen_{etudiant.id}") or 0
            )


            Note.objects.update_or_create(
                etudiant=etudiant,
                matiere_id=matiere_id,
                semestre=semestre,
                defaults={
                    "saisie": saisie,
                    "cc": cc,
                    "devoir": devoir,
                    "examen": examen,
                }
            )


        messages.success(
            request,
            "Les notes ont été enregistrées avec succès."
        )


        return redirect(
            f"{request.path}?classe={classe_id}&matiere={matiere_id}&semestre={semestre}"
        )



    # ==========================
    # AFFICHAGE
    # ==========================

    if classe_id and matiere_id and semestre:


        classe = Classe.objects.get(
            id=classe_id
        )


        etudiants = Etudiant.objects.filter(
            classe=classe
        )


        notes = Note.objects.filter(
            etudiant__in=etudiants,
            matiere_id=matiere_id,
            semestre=semestre
        )


        for note in notes:

            notes_existantes[note.etudiant_id] = note



        # envoyer la note directement dans le template
        for etudiant in etudiants:

            etudiant.note_existante = notes_existantes.get(
                etudiant.id
            )



    context = {

        "classes": classes,

        "matieres": matieres,

        "etudiants": etudiants,

        "notes_existantes": notes_existantes,

    }


    return render(
        request,
        "notes/saisie_groupee.html",
        context
    )
import unicodedata
    

import re



# ==============================================================
# NORMALISATION GÉNÉRALE
# ==============================================================

def normaliser_texte(texte):
    """
    Normalise un texte pour les comparaisons :

    - minuscules
    - suppression des accents
    - apostrophes uniformisées
    - espaces multiples supprimés
    """

    if texte is None:
        return ""

    texte = str(texte).strip().lower()

    # Uniformiser les apostrophes
    texte = texte.replace("’", "'")
    texte = texte.replace("`", "'")
    texte = texte.replace("´", "'")

    # Supprimer les accents
    texte = unicodedata.normalize(
        "NFD",
        texte
    )

    texte = "".join(
        caractere
        for caractere in texte
        if unicodedata.category(caractere) != "Mn"
    )

    # Remplacer les espaces multiples par un seul espace
    texte = " ".join(texte.split())

    return texte


# ==============================================================
# NORMALISATION FILIÈRE
# ==============================================================

def normaliser_filiere(texte):
    """
    Normalise une filière pour permettre les comparaisons
    même si le fichier Excel ne contient pas le code (IDA), (AD), etc.
    """

    texte = normaliser_texte(texte)

    # Supprimer les codes entre parenthèses
    # Exemple :
    # INFORMATIQUE ET DEVELOPPEMENT D'APPLICATIONS (IDA)
    # devient :
    # INFORMATIQUE ET DEVELOPPEMENT D'APPLICATIONS
    texte = re.sub(
        r"\s*\([^)]*\)",
        "",
        texte
    )

    # Uniformiser les apostrophes restantes
    texte = texte.replace("'", "")

    # Supprimer les espaces multiples
    texte = " ".join(texte.split())

    return texte


# ==============================================================
# IMPORT DES ÉTUDIANTS EXCEL
# ==============================================================

from datetime import datetime, date

# ----------------------------------------------------------
# Conversion flexible des dates
# ----------------------------------------------------------

def convertir_date(valeur):

    if valeur is None:
        return None

    # Déjà une date Python
    if isinstance(valeur, date) and not isinstance(valeur, datetime):
        return valeur

    # Date Excel (datetime)
    if isinstance(valeur, datetime):
        return valeur.date()

    # Nombre Excel (parfois 45000, etc.)
    if isinstance(valeur, (int, float)):
        try:
            return datetime.fromordinal(
                datetime(1899, 12, 30).toordinal() + int(valeur)
            ).date()
        except Exception:
            return None

    # Texte
    if isinstance(valeur, str):

        texte = valeur.strip()

        formats = [
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%d.%m.%Y",
            "%d %m %Y",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(texte, fmt).date()
            except ValueError:
                continue

    return None


# ----------------------------------------------------------
# Importation des étudiants
# ----------------------------------------------------------

def import_etudiants_excelDDD(request):

    if request.method != "POST":
        return render(request, "import_etudiants_excel.html")

    fichier = request.FILES.get("excel_file")

    if not fichier:
        messages.error(
            request,
            "Veuillez sélectionner un fichier Excel."
        )
        return redirect("import_etudiants_excel")

    if not fichier.name.lower().endswith(".xlsx"):
        messages.error(
            request,
            "Format incorrect. Veuillez importer uniquement un fichier .xlsx."
        )
        return redirect("import_etudiants_excel")

    try:
        wb = load_workbook(fichier, data_only=True)
    except Exception as e:
        messages.error(
            request,
            f"Impossible de lire le fichier Excel : {e}"
        )
        return redirect("import_etudiants_excel")

    ws = wb.active

    compteur_creation = 0
    compteur_modification = 0
    erreurs = []

    for ligne, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):

        try:

            if not row or all(
                value is None or str(value).strip() == ""
                for value in row
            ):
                continue

            matricule = str(row[0]).strip() if row[0] else ""
            nom = str(row[1]).strip() if row[1] else ""
            prenoms = str(row[2]).strip() if row[2] else ""

            # -----------------------------
            # Date flexible
            # -----------------------------
            date_naissance = convertir_date(row[3])

            if not date_naissance:

                erreurs.append(
                    f"Ligne {ligne}: date de naissance invalide."
                )

                continue

            lieu_naissance = str(row[4]).strip() if row[4] else ""
            sexe = str(row[5]).strip().upper() if row[5] else ""
            telephone = str(row[6]).strip() if row[6] else ""
            email = str(row[7]).strip().lower() if row[7] else ""
            classe_nom = str(row[8]).strip() if row[8] else ""
            filiere_nom = str(row[9]).strip() if row[9] else ""

            # -----------------------------
            # Recherche classe
            # -----------------------------
            classe = Classe.objects.filter(
                nom__iexact=classe_nom
            ).first()

            if not classe:

                erreurs.append(
                    f"Ligne {ligne}: classe '{classe_nom}' introuvable."
                )

                continue

            # -----------------------------
            # Recherche filière
            # -----------------------------
            filiere = Filierebts.objects.filter(
                nom__iexact=filiere_nom
            ).first()

            if not filiere:

                erreurs.append(
                    f"Ligne {ligne}: filière '{filiere_nom}' introuvable."
                )

                continue

            # -----------------------------
            # Étudiant existant ?
            # -----------------------------
            etudiant = Etudiant.objects.filter(
                matricule__iexact=matricule
            ).first()

            if etudiant:

                etudiant.nom = nom
                etudiant.prenoms = prenoms
                etudiant.date_naissance = date_naissance
                etudiant.lieu_naissance = lieu_naissance
                etudiant.sexe = sexe
                etudiant.telephone = telephone
                etudiant.email = email
                etudiant.classe = classe
                etudiant.filiere_bts = filiere
                etudiant.save()

                compteur_modification += 1

            else:

                Etudiant.objects.create(
                    matricule=matricule,
                    nom=nom,
                    prenoms=prenoms,
                    date_naissance=date_naissance,
                    lieu_naissance=lieu_naissance,
                    sexe=sexe,
                    telephone=telephone,
                    email=email,
                    classe=classe,
                    filiere_bts=filiere,
                )

                compteur_creation += 1

        except Exception as e:

            erreurs.append(
                f"Ligne {ligne}: erreur inattendue : {e}"
            )

    # ----------------------------------------------------------
    # Messages de résultat
    # ----------------------------------------------------------

    if compteur_creation:
        messages.success(
            request,
            f"{compteur_creation} étudiant(s) créé(s) avec succès."
        )

    if compteur_modification:
        messages.success(
            request,
            f"{compteur_modification} étudiant(s) mis à jour avec succès."
        )

    if erreurs:

        messages.warning(
            request,
            f"{len(erreurs)} ligne(s) n'ont pas été importée(s)."
        )

        for erreur in erreurs:
            messages.warning(request, erreur)

    return redirect("etudiant_list")

def import_etudiants_excel(request):

    if request.method != "POST":
        return render(request, "import_etudiants_excel.html")

    fichier = request.FILES.get("excel_file")

    if not fichier:
        messages.error(
            request,
            "Veuillez sélectionner un fichier Excel."
        )
        return redirect("import_etudiants_excel")

    if not fichier.name.lower().endswith(".xlsx"):
        messages.error(
            request,
            "Format incorrect. Veuillez importer uniquement un fichier .xlsx."
        )
        return redirect("import_etudiants_excel")

    try:
        wb = load_workbook(fichier, data_only=True)
    except Exception as e:
        messages.error(
            request,
            f"Impossible de lire le fichier Excel : {e}"
        )
        return redirect("import_etudiants_excel")

    ws = wb.active

    compteur_creation = 0
    compteur_modification = 0
    erreurs = []

    for ligne, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):

        try:

            # ------------------------------------------------------
            # IGNORER LES LIGNES VIDES
            # ------------------------------------------------------

            if not row or all(
                value is None or str(value).strip() == ""
                for value in row
            ):
                continue

            # ------------------------------------------------------
            # LECTURE DES COLONNES EXCEL
            # ------------------------------------------------------

            matricule = str(row[0]).strip() if row[0] else ""

            identifiant_permanent = (
                str(row[1]).strip()
                if row[1]
                else ""
            )

            nom = str(row[2]).strip() if row[2] else ""

            prenoms = (
                str(row[3]).strip()
                if row[3]
                else ""
            )

            # ------------------------------------------------------
            # DATE DE NAISSANCE
            # ------------------------------------------------------

            date_naissance = convertir_date(row[4])

            if not date_naissance:

                erreurs.append(
                    f"Ligne {ligne}: date de naissance invalide."
                )

                continue

            # ------------------------------------------------------
            # AUTRES INFORMATIONS
            # ------------------------------------------------------

            lieu_naissance = (
                str(row[5]).strip()
                if row[5]
                else ""
            )

            sexe = (
                str(row[6]).strip().upper()
                if row[6]
                else ""
            )

            telephone = (
                str(row[7]).strip()
                if row[7]
                else ""
            )

            email = (
                str(row[8]).strip().lower()
                if row[8]
                else ""
            )

            classe_nom = (
                str(row[9]).strip()
                if row[9]
                else ""
            )

            filiere_nom = (
                str(row[10]).strip()
                if row[10]
                else ""
            )

            # ------------------------------------------------------
            # VÉRIFICATION MATRICULE
            # ------------------------------------------------------

            if not matricule:

                erreurs.append(
                    f"Ligne {ligne}: matricule obligatoire."
                )

                continue

            # ------------------------------------------------------
            # VÉRIFICATION IP
            # ------------------------------------------------------

            if not identifiant_permanent:

                erreurs.append(
                    f"Ligne {ligne}: identifiant permanent (IP) obligatoire."
                )

                continue

            # ------------------------------------------------------
            # RECHERCHE CLASSE
            # ------------------------------------------------------

            classe = Classe.objects.filter(
                nom__iexact=classe_nom
            ).first()

            if not classe:

                erreurs.append(
                    f"Ligne {ligne}: classe '{classe_nom}' introuvable."
                )

                continue

            # ------------------------------------------------------
            # RECHERCHE FILIÈRE
            # ------------------------------------------------------

            filiere = Filierebts.objects.filter(
                nom__iexact=filiere_nom
            ).first()

            if not filiere:

                erreurs.append(
                    f"Ligne {ligne}: filière '{filiere_nom}' introuvable."
                )

                continue

            # ------------------------------------------------------
            # RECHERCHE ÉTUDIANT PAR MATRICULE
            # ------------------------------------------------------

            etudiant = Etudiant.objects.filter(
                matricule__iexact=matricule
            ).first()

            if etudiant:

                # --------------------------------------------------
                # MISE À JOUR
                # --------------------------------------------------

                etudiant.identifiant_permanent = identifiant_permanent
                etudiant.nom = nom
                etudiant.prenoms = prenoms
                etudiant.date_naissance = date_naissance
                etudiant.lieu_naissance = lieu_naissance
                etudiant.sexe = sexe
                etudiant.telephone = telephone
                etudiant.email = email
                etudiant.classe = classe
                etudiant.filiere_bts = filiere

                etudiant.save()

                compteur_modification += 1

            else:

                # --------------------------------------------------
                # CRÉATION
                # --------------------------------------------------

                Etudiant.objects.create(

                    matricule=matricule,

                    identifiant_permanent=identifiant_permanent,

                    nom=nom,

                    prenoms=prenoms,

                    date_naissance=date_naissance,

                    lieu_naissance=lieu_naissance,

                    sexe=sexe,

                    telephone=telephone,

                    email=email,

                    classe=classe,

                    filiere_bts=filiere,
                )

                compteur_creation += 1

        except Exception as e:

            erreurs.append(
                f"Ligne {ligne}: erreur inattendue : {e}"
            )

    # ----------------------------------------------------------
    # MESSAGES DE RÉSULTAT
    # ----------------------------------------------------------

    if compteur_creation:

        messages.success(
            request,
            f"{compteur_creation} étudiant(s) créé(s) avec succès."
        )

    if compteur_modification:

        messages.success(
            request,
            f"{compteur_modification} étudiant(s) mis à jour avec succès."
        )

    if erreurs:

        messages.warning(
            request,
            f"{len(erreurs)} ligne(s) n'ont pas été importée(s)."
        )

        for erreur in erreurs:

            messages.warning(
                request,
                erreur
            )

    return redirect("etudiant_list")


def matieres_par_classe(request):
    classe_id = request.GET.get("classe")

    if not classe_id:
        return JsonResponse([], safe=False)

    try:
        classe = Classe.objects.select_related("filiere_bts").get(pk=classe_id)
    except Classe.DoesNotExist:
        return JsonResponse([], safe=False)

    matieres = Matiere.objects.filter(
        filiere_bts=classe.filiere_bts
    ).order_by("libelle")

    data = [
        {
            "id": m.id,
            "code": m.code,
            "libelle": m.libelle,
        }
        for m in matieres
    ]

    return JsonResponse(data, safe=False)

def grande_unite_list(request):

    filiere_id = request.GET.get("filiere_bts")
    q = request.GET.get("q", "").strip()

    grandes_unites = GrandeUnite.objects.select_related(
        "filiere_bts"
    ).all()

    if filiere_id:
        grandes_unites = grandes_unites.filter(
            filiere_bts_id=filiere_id
        )

    if q:
        grandes_unites = grandes_unites.filter(
            models.Q(code__icontains=q) |
            models.Q(libelle__icontains=q)
        )

    filiere_list = Filierebts.objects.all().order_by("nom")

    context = {
        "grandes_unites": grandes_unites,
        "filiere_list": filiere_list,
        "filiere_selectionnee": filiere_id,
        "total_grandes_unites": grandes_unites.count(),
    }

    return render(
        request,
        "matieres/grande_unite_list.html",
        context
    )

def grande_unite_add(request):

    if request.method == "POST":

        code = request.POST.get("code", "").strip()
        libelle = request.POST.get("libelle", "").strip()
        ordre = request.POST.get("ordre", "1")
        filiere_bts = request.POST.get("filiere_bts")
        description = request.POST.get("description", "").strip()

        if not code or not libelle or not filiere_bts:

            messages.error(
                request,
                "Veuillez remplir tous les champs obligatoires."
            )

            return redirect("grande_unite_add")

        try:
            ordre = int(ordre)
        except (TypeError, ValueError):
            ordre = 1

        GrandeUnite.objects.create(
            code=code,
            libelle=libelle,
            ordre=ordre,
            filiere_bts_id=filiere_bts,
            description=description
        )

        messages.success(
            request,
            "Grande unité créée avec succès."
        )

        return redirect("grande_unite_list")

    filiere_list = Filierebts.objects.all().order_by("nom")

    context = {
        "filiere_list": filiere_list
    }

    return render(
        request,
        "matieres/grande_unite_form.html",
        context
    )

def grande_unite_edit(request, pk):

    grande_unite = get_object_or_404(
        GrandeUnite,
        pk=pk
    )

    if request.method == "POST":

        code = request.POST.get("code", "").strip()
        libelle = request.POST.get("libelle", "").strip()
        ordre = request.POST.get("ordre", "1")
        filiere_bts = request.POST.get("filiere_bts")
        description = request.POST.get("description", "").strip()

        if not code or not libelle or not filiere_bts:

            messages.error(
                request,
                "Veuillez remplir tous les champs obligatoires."
            )

            return redirect(
                "grande_unite_edit",
                pk=pk
            )

        try:
            ordre = int(ordre)
        except (TypeError, ValueError):
            ordre = 1

        grande_unite.code = code
        grande_unite.libelle = libelle
        grande_unite.ordre = ordre
        grande_unite.filiere_bts_id = filiere_bts
        grande_unite.description = description

        grande_unite.save()

        messages.success(
            request,
            "Grande unité modifiée avec succès."
        )

        return redirect("grande_unite_list")

    filiere_list = Filierebts.objects.all().order_by("nom")

    context = {
        "grande_unite": grande_unite,
        "filiere_list": filiere_list
    }

    return render(
        request,
        "matieres/grande_unite_form.html",
        context
    )

def grande_unite_delete(request, pk):

    grande_unite = get_object_or_404(
        GrandeUnite,
        pk=pk
    )

    if request.method == "POST":

        grande_unite.delete()

        messages.success(
            request,
            "Grande unité supprimée avec succès."
        )

        return redirect("grande_unite_list")

    return render(
        request,
        "matieres/grande_unite_confirm_delete.html",
        {
            "grande_unite": grande_unite
        }
    )

def export_etudiants_pdf(request):

    # =====================================================
    # RÉCUPÉRATION DES FILTRES
    # =====================================================

    q = request.GET.get("q", "").strip()
    niveau = request.GET.get("niveau", "").strip()
    filiere_bts = request.GET.get("filiere_bts", "").strip()

    # =====================================================
    # REQUÊTE ÉTUDIANTS
    # =====================================================

    etudiants = Etudiant.objects.select_related(
        "classe",
        "filiere_bts"
    ).all()

    if q:
        etudiants = etudiants.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(matricule__icontains=q) |
            Q(identifiant_permanent__icontains=q)
        )

    if niveau:
        etudiants = etudiants.filter(
            classe__nom__icontains=niveau
        )

    if filiere_bts:
        etudiants = etudiants.filter(
            filiere_bts_id=filiere_bts
        )

    etudiants = etudiants.order_by(
        "nom",
        "prenoms"
    )

    # =====================================================
    # RÉPONSE PDF
    # =====================================================

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="liste_etudiants_BTS.pdf"'
    )

    # =====================================================
    # DOCUMENT A4 PAYSAGE
    # =====================================================

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=0.7 * cm,
        rightMargin=0.7 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )

    elements = []

    # =====================================================
    # STYLES
    # =====================================================

    styles = getSampleStyleSheet()

    titre = ParagraphStyle(
        "TitreEtudiants",
        parent=styles["Heading1"],
        alignment=TA_CENTER,
        fontSize=16,
        leading=19,
        spaceAfter=4,
        textColor=colors.HexColor("#071d35"),
        fontName="Helvetica-Bold",
    )

    sous_titre = ParagraphStyle(
        "SousTitreEtudiants",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=8,
        leading=11,
        spaceAfter=10,
        textColor=colors.HexColor("#555555"),
    )

    # Style des cellules
    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=8,
        alignment=TA_LEFT,
    )

    cell_center_style = ParagraphStyle(
        "CellCenterStyle",
        parent=cell_style,
        alignment=TA_CENTER,
    )

    header_style = ParagraphStyle(
        "HeaderStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=8,
        alignment=TA_CENTER,
        textColor=colors.white,
    )

    # =====================================================
    # EN-TÊTE
    # =====================================================

    elements.append(
        Paragraph(
            "GROUPE EXPERT MÉTIER",
            titre
        )
    )

    elements.append(
        Paragraph(
            "LISTE DES ÉTUDIANTS BTS",
            titre
        )
    )

    # =====================================================
    # INFORMATIONS FILTRES
    # =====================================================

    informations = []

    if filiere_bts:

        filiere = Filierebts.objects.filter(
            id=filiere_bts
        ).first()

        if filiere:
            informations.append(
                f"<b>Filière :</b> {filiere.nom}"
            )

    if niveau:
        informations.append(
            f"<b>Niveau :</b> {niveau}"
        )

    if q:
        informations.append(
            f"<b>Recherche :</b> {q}"
        )

    informations.append(
        f"<b>Total :</b> {etudiants.count()} étudiant(s)"
    )

    elements.append(
        Paragraph(
            " &nbsp; | &nbsp; ".join(informations),
            sous_titre
        )
    )

    # =====================================================
    # TABLEAU
    # =====================================================

    data = [
        [
            Paragraph("N°", header_style),
            Paragraph("IDENTIFIANT<br/>PERMANENT (IP)", header_style),
            Paragraph("MATRICULE", header_style),
            Paragraph("NOM", header_style),
            Paragraph("PRÉNOMS", header_style),
            Paragraph("SEXE", header_style),
            Paragraph("CLASSE", header_style),
            Paragraph("FILIÈRE", header_style),
        ]
    ]

    # =====================================================
    # LIGNES ÉTUDIANTS
    # =====================================================

    for index, e in enumerate(etudiants, start=1):

        # -------------------------------------------------
        # SEXE
        # -------------------------------------------------

        if e.sexe:
            sexe = e.get_sexe_display()
        else:
            sexe = "—"

        # -------------------------------------------------
        # IP
        # -------------------------------------------------

        ip = (
            e.identifiant_permanent
            if e.identifiant_permanent
            else "—"
        )

        # -------------------------------------------------
        # CLASSE
        # -------------------------------------------------

        classe = (
            str(e.classe)
            if e.classe
            else "—"
        )

        # -------------------------------------------------
        # FILIÈRE
        # -------------------------------------------------

        filiere = (
            e.filiere_bts.nom
            if e.filiere_bts
            else "—"
        )

        data.append([
            Paragraph(
                str(index),
                cell_center_style
            ),

            Paragraph(
                str(ip),
                cell_center_style
            ),

            Paragraph(
                str(e.matricule or "—"),
                cell_center_style
            ),

            Paragraph(
                str(e.nom or "—"),
                cell_style
            ),

            Paragraph(
                str(e.prenoms or "—"),
                cell_style
            ),

            Paragraph(
                str(sexe),
                cell_center_style
            ),

            Paragraph(
                classe,
                cell_style
            ),

            Paragraph(
                filiere,
                cell_style
            ),
        ])

    # =====================================================
    # LARGEUR DES COLONNES
    # =====================================================

    # A4 paysage = 29.7 cm
    # Marges 0.7 + 0.7 = 1.4 cm
    # Largeur disponible ≈ 28.3 cm

    col_widths = [
        0.8 * cm,   # N°
        3.5 * cm,   # IP
        2.7 * cm,   # Matricule
        3.2 * cm,   # Nom
        4.4 * cm,   # Prénoms
        2.0 * cm,   # Sexe
        3.8 * cm,   # Classe
        5.9 * cm,   # Filière
    ]

    # Total = 26.3 cm environ
    # donc aucune colonne ne dépasse la page.

    table = Table(
        data,
        colWidths=col_widths,
        repeatRows=1,
        splitByRow=1,
    )

    # =====================================================
    # STYLE TABLEAU
    # =====================================================

    table.setStyle(
        TableStyle([

            # -------------------------------------------------
            # EN-TÊTE
            # -------------------------------------------------

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#071d35")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, 0),
                "CENTER"
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            # -------------------------------------------------
            # CORPS
            # -------------------------------------------------

            (
                "FONTNAME",
                (0, 1),
                (-1, -1),
                "Helvetica"
            ),

            (
                "FONTSIZE",
                (0, 1),
                (-1, -1),
                7
            ),

            # -------------------------------------------------
            # BORDURES
            # -------------------------------------------------

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.4,
                colors.HexColor("#b8c0ca")
            ),

            (
                "BOX",
                (0, 0),
                (-1, -1),
                0.8,
                colors.HexColor("#071d35")
            ),

            # -------------------------------------------------
            # ALIGNEMENT
            # -------------------------------------------------

            (
                "ALIGN",
                (0, 0),
                (2, -1),
                "CENTER"
            ),

            (
                "ALIGN",
                (5, 0),
                (5, -1),
                "CENTER"
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            # -------------------------------------------------
            # ALTERNANCE DES LIGNES
            # -------------------------------------------------

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f5f7fb")
                ]
            ),

            # -------------------------------------------------
            # ESPACEMENT
            # -------------------------------------------------

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                4
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                4
            ),

            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                3
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                3
            ),
        ])
    )

    elements.append(table)

    elements.append(
        Spacer(1, 12)
    )

    # =====================================================
    # FOOTER
    # =====================================================

    footer_style = ParagraphStyle(
        "FooterEtudiants",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=6.5,
        leading=8,
        textColor=colors.grey,
    )

    elements.append(
        Paragraph(
            "Document généré automatiquement par le système de gestion académique GEM.",
            footer_style
        )
    )

    # =====================================================
    # GÉNÉRATION
    # =====================================================

    doc.build(elements)

    return response